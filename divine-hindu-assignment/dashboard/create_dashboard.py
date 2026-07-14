import os
import json
import subprocess
import sys

# Step 1: Programmatically ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl library for Excel generation...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_command_center():
    wb = openpyxl.Workbook()
    
    # Define styles
    primary_color = "E65100"  # Saffron
    accent_color = "FFB300"   # Gold
    dark_navy = "1A1A2E"      # Navy Blue for main headers
    light_orange = "FFF3E0"   # Zebra background/light highlight
    
    font_family = "Segoe UI"
    
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    dark_header_fill = PatternFill(start_color=dark_navy, end_color=dark_navy, fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    card_title_font = Font(name=font_family, size=11, bold=True, color=primary_color)
    card_value_font = Font(name=font_family, size=20, bold=True, color="1A1A2E")
    
    # -------------------------------------------------------------------------
    # TAB 1: Master Tracker
    # -------------------------------------------------------------------------
    ws_master = wb.active
    ws_master.title = "Master Tracker"
    
    master_headers = [
        "Timestamp", "Platform", "Post ID", "Comment ID", "Username", "Comment Text",
        "AI Category", "AI Confidence", "Priority", "Action", "Suggested Reply",
        "Action Taken", "Assigned To", "Status", "Response Time (min)", "Notes"
    ]
    
    for col_idx, header in enumerate(master_headers, 1):
        cell = ws_master.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
    
    # -------------------------------------------------------------------------
    # TAB 2: DM Tracker
    # -------------------------------------------------------------------------
    ws_dm = wb.create_sheet(title="DM Tracker")
    dm_headers = [
        "Timestamp", "Platform", "Conversation ID", "Customer Username", "Message Text",
        "AI Category", "Priority", "Route To", "Suggested Reply (Hindi)",
        "Suggested Reply (English)", "Urgency", "Status", "Handled By", "Resolution Time (min)"
    ]
    for col_idx, header in enumerate(dm_headers, 1):
        cell = ws_dm.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
        
    # -------------------------------------------------------------------------
    # TAB 3: Daily Dashboard
    # -------------------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Daily Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells("A1:H2")
    title_cell = ws_dash["A1"]
    title_cell.value = " 📊 Divine Hindu — Social Media Intelligence Dashboard"
    title_cell.font = title_font
    title_cell.fill = dark_header_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Subtitle for Date
    ws_dash["A3"].value = "Today:"
    ws_dash["A3"].font = Font(name=font_family, size=10, bold=True)
    ws_dash["B3"].value = '=TEXT(TODAY(), "DD MMM YYYY")'
    ws_dash["B3"].font = Font(name=font_family, size=10, italic=True)
    
    # KPI 1: Total Comments Today
    ws_dash.merge_cells("B4:C4")
    ws_dash["B4"].value = "TOTAL COMMENTS TODAY"
    ws_dash["B4"].font = Font(name=font_family, size=9, bold=True, color="666666")
    ws_dash["B4"].alignment = Alignment(horizontal="center")
    ws_dash.merge_cells("B5:C5")
    ws_dash["B5"].value = "=COUNTIF('Master Tracker'!A:A, \">=\"&TODAY())"
    ws_dash["B5"].font = card_value_font
    ws_dash["B5"].alignment = Alignment(horizontal="center", vertical="center")
    
    # KPI 2: Avg Response Time
    ws_dash.merge_cells("E4:F4")
    ws_dash["E4"].value = "AVG RESPONSE TIME"
    ws_dash["E4"].font = Font(name=font_family, size=9, bold=True, color="666666")
    ws_dash["E4"].alignment = Alignment(horizontal="center")
    ws_dash.merge_cells("E5:F5")
    ws_dash["E5"].value = "=IFERROR(AVERAGEIFS('Master Tracker'!O:O, 'Master Tracker'!A:A, \">=\"&TODAY()), 0)"
    ws_dash["E5"].font = card_value_font
    ws_dash["E5"].alignment = Alignment(horizontal="center", vertical="center")
    
    # KPI 3: Auto-Resolved %
    ws_dash.merge_cells("H4:I4")
    ws_dash["H4"].value = "AUTO-RESOLVED %"
    ws_dash["H4"].font = Font(name=font_family, size=9, bold=True, color="666666")
    ws_dash["H4"].alignment = Alignment(horizontal="center")
    ws_dash.merge_cells("H5:I5")
    ws_dash["H5"].value = "=IFERROR(COUNTIFS('Master Tracker'!A:A, \">=\"&TODAY(), 'Master Tracker'!J:J, \"AUTO_REPLY\") / B5 * 100, 100)"
    ws_dash["H5"].font = card_value_font
    ws_dash["H5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash["H5"].number_format = '0.0"%"'
    
    # Card borders
    card_fill = PatternFill(start_color=light_orange, end_color=light_orange, fill_type="solid")
    for row in [4, 5]:
        for col in [2, 3, 5, 6, 8, 9]:
            cell = ws_dash.cell(row=row, column=col)
            cell.fill = card_fill
            cell.border = thin_border
            
    # Category Breakdown Table
    ws_dash["B7"].value = "Category"
    ws_dash["B7"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws_dash["B7"].fill = header_fill
    ws_dash["B7"].border = thin_border
    
    ws_dash["C7"].value = "Volume Today"
    ws_dash["C7"].font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    ws_dash["C7"].fill = header_fill
    ws_dash["C7"].border = thin_border
    ws_dash["C7"].alignment = Alignment(horizontal="right")
    
    categories = [
        "PRODUCT_INQUIRY", "PRICE_QUESTION", "COMPLAINT", "ORDER_QUERY",
        "SPAM", "IRRELEVANT", "POSITIVE_FEEDBACK", "COLLABORATION_REQUEST"
    ]
    
    for idx, cat in enumerate(categories, 8):
        # Category Name
        cell_name = ws_dash.cell(row=idx, column=2, value=cat.replace("_", " ").title())
        cell_name.font = Font(name=font_family, size=10)
        cell_name.border = thin_border
        
        # Category Count Formula
        cell_formula = ws_dash.cell(row=idx, column=3, value=f"=COUNTIFS('Master Tracker'!A:A, \">=\"&TODAY(), 'Master Tracker'!G:G, \"{cat}\")")
        cell_formula.font = Font(name=font_family, size=10, bold=True)
        cell_formula.border = thin_border
        cell_formula.alignment = Alignment(horizontal="right")

    # -------------------------------------------------------------------------
    # TAB 4: Reply Templates (Pre-populated from JSON)
    # -------------------------------------------------------------------------
    ws_templates = wb.create_sheet(title="Reply Templates")
    templates_headers = [
        "Template ID", "Category", "Sub-Category", "Template (Hindi)",
        "Template (English)", "Variables", "Tone", "Platform"
    ]
    for col_idx, header in enumerate(templates_headers, 1):
        cell = ws_templates.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Load from reply-templates.json
    script_dir = os.path.dirname(os.path.abspath(__file__))
    templates_json_path = os.path.join(script_dir, "..", "templates", "reply-templates.json")
    
    if os.path.exists(templates_json_path):
        try:
            with open(templates_json_path, 'r', encoding='utf-8') as f:
                templates_data = json.load(f)
            
            for row_idx, temp in enumerate(templates_data, 2):
                ws_templates.cell(row=row_idx, column=1, value=temp.get("id", ""))
                ws_templates.cell(row=row_idx, column=2, value=temp.get("category", ""))
                ws_templates.cell(row=row_idx, column=3, value=temp.get("subcategory", ""))
                ws_templates.cell(row=row_idx, column=4, value=temp.get("template_hindi", ""))
                ws_templates.cell(row=row_idx, column=5, value=temp.get("template_english", ""))
                ws_templates.cell(row=row_idx, column=6, value=", ".join(temp.get("variables", [])))
                ws_templates.cell(row=row_idx, column=7, value=temp.get("tone", ""))
                ws_templates.cell(row=row_idx, column=8, value=temp.get("platform", ""))
                
                # Apply styles to templates rows
                for c in range(1, 9):
                    cell = ws_templates.cell(row=row_idx, column=c)
                    cell.font = Font(name=font_family, size=9.5)
                    cell.border = thin_border
                    if c in [4, 5]: # Wrap reply texts
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
        except Exception as e:
            print(f"Warning: Could not populate templates sheet: {str(e)}")
            
    # -------------------------------------------------------------------------
    # TAB 5: Daily Reports
    # -------------------------------------------------------------------------
    ws_reports = wb.create_sheet(title="Daily Reports")
    reports_headers = [
        "Date", "Total Comments", "Total DMs", "Avg Response Time",
        "Auto-Resolved %", "Hot Leads", "Complaints", "Spam Cleaned", "AI Summary"
    ]
    for col_idx, header in enumerate(reports_headers, 1):
        cell = ws_reports.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # -------------------------------------------------------------------------
    # TAB 6: Config (Pre-populated)
    # -------------------------------------------------------------------------
    ws_config = wb.create_sheet(title="Config")
    
    config_headers = ["Configuration Key", "Value", "Description"]
    for col_idx, header in enumerate(config_headers, 1):
        cell = ws_config.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = dark_header_fill
        cell.border = thin_border
        
    config_rows = [
        ["Last Processed Comment ID (IG)", "", "Automatically updated by n8n"],
        ["Last Processed Comment ID (FB)", "", "Automatically updated by n8n"],
        ["Last Processed DM Timestamp", "", "Automatically updated by n8n"],
        ["Team Members", "Aarav, Priya, Rohan", "Comma-separated list of handlers"],
        ["Alert Email", "manager@divinehindu.in", "Recipients of escalation alerts"],
        ["Support Phone", "+91-9876543210", "Displayed in customer replies"],
        ["Min Bulk Quantity", 50, "Minimum units threshold for bulk rates"],
        ["Free Shipping Threshold", 999, "Min order amount in rupees for free shipping"]
    ]
    
    for r_idx, config_item in enumerate(config_rows, 2):
        for c_idx, val in enumerate(config_item, 1):
            cell = ws_config.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name=font_family, size=10, bold=(c_idx==1))
            cell.border = thin_border
            if c_idx == 1:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # -------------------------------------------------------------------------
    # Auto-adjust column widths across all sheets
    # -------------------------------------------------------------------------
    for sheet in wb.worksheets:
        if sheet.title == "Daily Dashboard":
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 18
            sheet.column_dimensions['E'].width = 18
            sheet.column_dimensions['F'].width = 18
            sheet.column_dimensions['G'].width = 18
            sheet.column_dimensions['H'].width = 18
            sheet.column_dimensions['I'].width = 18
            continue
            
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # For wrapped text fields, don't make columns extremely wide
            if sheet.title == "Reply Templates" and col_letter in ['D', 'E']:
                sheet.column_dimensions[col_letter].width = 45
                continue
                
            for cell in col:
                val_str = str(cell.value or '')
                # If formula, estimate length by header length
                if val_str.startswith('='):
                    val_str = str(col[0].value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save spreadsheet
    output_filename = os.path.join(script_dir, "Divine_Hindu_Social_Media_Command_Center.xlsx")
    wb.save(output_filename)
    print(f"Spreadsheet generated successfully: {output_filename}")
    return output_filename

if __name__ == "__main__":
    create_command_center()
